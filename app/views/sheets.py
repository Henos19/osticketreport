from flask import Blueprint
from flask import send_file

from io import BytesIO
from app.models import Ticket

import pandas as pd
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter


sheets = Blueprint('sheets', __name__)


@sheets.route('/excel/tickets.xlsx')
def export_tickets():
    tickets = Ticket.query.all()
    data = [{
        'Data de Criação': tck.created,
        'Nome': tck.tk_name[0].name,
        'Conteúdo': tck.tk_title[0].content,
        'Ticket': tck.number,
        'Data de Finalzação': tck.closed,
    } for tck in tickets]

    df = pd.DataFrame(data)
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Tickets')
        sheet = writer.sheets['Tickets']

        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(bold=True)


        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')


        for row in sheet.iter_rows(min_row=2, min_col=0, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

        # for row in sheet.iter_rows(min_row=1, min_col=3, max_col=3):
        #     for cell in row:
        #         cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2


    output.seek(0)
    return send_file(
        output,
        download_name="tickets.xlsx",
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


